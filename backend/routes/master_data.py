from flask import Blueprint, request, jsonify
from db import get_db, rows_to_list, row_to_dict

master_bp = Blueprint('master', __name__)

def nd(v):
    if v == '' or v is None:
        return None
    return v

# ── GRADES ────────────────────────────────────────────────────────

@master_bp.route('/api/grades', methods=['POST'])
def create_grade():
    d = request.json
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            INSERT INTO grades (
                grade_code, c_min, c_max, mn_min, mn_max, p_max,
                s_min, s_max, si_min, si_max, ni_min, ni_max,
                mo_min, mo_max, cr_min, cr_max, n_max, cu_min, cu_max
            ) VALUES (
                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s
            )
        """, (
            d.get('grade_code'),
            nd(d.get('c_min')),   nd(d.get('c_max')),
            nd(d.get('mn_min')),  nd(d.get('mn_max')),
            nd(d.get('p_max')),
            nd(d.get('s_min')),   nd(d.get('s_max')),
            nd(d.get('si_min')),  nd(d.get('si_max')),
            nd(d.get('ni_min')),  nd(d.get('ni_max')),
            nd(d.get('mo_min')),  nd(d.get('mo_max')),
            nd(d.get('cr_min')),  nd(d.get('cr_max')),
            nd(d.get('n_max')),
            nd(d.get('cu_min')),  nd(d.get('cu_max')),
        ))
        db.commit()
        return jsonify({'success': True, 'id': cur.lastrowid})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@master_bp.route('/api/grades/<int:gid>', methods=['PUT'])
def update_grade(gid):
    d = request.json
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            UPDATE grades SET
                grade_code=%s, c_min=%s, c_max=%s,
                mn_min=%s, mn_max=%s, p_max=%s,
                s_min=%s, s_max=%s, si_min=%s, si_max=%s,
                ni_min=%s, ni_max=%s, mo_min=%s, mo_max=%s,
                cr_min=%s, cr_max=%s, n_max=%s,
                cu_min=%s, cu_max=%s
            WHERE id=%s
        """, (
            d.get('grade_code'),
            nd(d.get('c_min')),   nd(d.get('c_max')),
            nd(d.get('mn_min')),  nd(d.get('mn_max')),
            nd(d.get('p_max')),
            nd(d.get('s_min')),   nd(d.get('s_max')),
            nd(d.get('si_min')),  nd(d.get('si_max')),
            nd(d.get('ni_min')),  nd(d.get('ni_max')),
            nd(d.get('mo_min')),  nd(d.get('mo_max')),
            nd(d.get('cr_min')),  nd(d.get('cr_max')),
            nd(d.get('n_max')),
            nd(d.get('cu_min')),  nd(d.get('cu_max')),
            gid,
        ))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

# ── CUSTOMERS ─────────────────────────────────────────────────────

@master_bp.route('/api/customers', methods=['POST'])
def create_customer():
    d = request.json
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            INSERT INTO customers (
                customer_name, short_code, country, address,
                contact_person, email, phone, payment_terms, incoterm
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            d.get('customer_name'),
            nd(d.get('short_code')),
            nd(d.get('country')),
            nd(d.get('address')),
            nd(d.get('contact_person')),
            nd(d.get('email')),
            nd(d.get('phone')),
            nd(d.get('payment_terms')),
            nd(d.get('incoterm')),
        ))
        db.commit()
        return jsonify({'success': True, 'id': cur.lastrowid})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@master_bp.route('/api/customers/<int:cid>', methods=['PUT'])
def update_customer(cid):
    d = request.json
    db = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            UPDATE customers SET
                customer_name=%s, short_code=%s, country=%s,
                address=%s, contact_person=%s, email=%s,
                phone=%s, payment_terms=%s, incoterm=%s
            WHERE id=%s
        """, (
            d.get('customer_name'),
            nd(d.get('short_code')),
            nd(d.get('country')),
            nd(d.get('address')),
            nd(d.get('contact_person')),
            nd(d.get('email')),
            nd(d.get('phone')),
            nd(d.get('payment_terms')),
            nd(d.get('incoterm')),
            cid,
        ))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


# ── BULK IMPORT ───────────────────────────────────────────────────

@master_bp.route('/api/customers/bulk-import', methods=['POST'])
def bulk_import_customers():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Please upload an .xlsx file'}), 400

    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(file.read()))
        ws = wb.active

        # Read headers from row 1
        headers = []
        for cell in ws[1]:
            if cell.value:
                # Map display labels to column names
                label_map = {
                    'Customer Name *': 'customer_name',
                    'Short Code': 'short_code',
                    'Country': 'country',
                    'Contact Person': 'contact_person',
                    'Email': 'email',
                    'Phone': 'phone',
                    'Payment Terms': 'payment_terms',
                    'Inco Term': 'incoterm',
                    'Address': 'address',
                }
                headers.append(label_map.get(cell.value, cell.value.lower().replace(' ', '_')))
            else:
                headers.append(None)

        db = get_db()
        cur = db.cursor()
        imported = 0
        skipped  = 0
        errors   = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue
            row_dict = dict(zip(headers, row))
            customer_name = str(row_dict.get('customer_name') or '').strip()
            if not customer_name:
                skipped += 1
                continue
            try:
                cur.execute("""
                    INSERT INTO customers (
                        customer_name, short_code, country, contact_person,
                        email, phone, payment_terms, incoterm, address, is_active
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,1)
                    ON DUPLICATE KEY UPDATE
                        short_code=VALUES(short_code),
                        country=VALUES(country),
                        contact_person=VALUES(contact_person),
                        email=VALUES(email),
                        phone=VALUES(phone),
                        payment_terms=VALUES(payment_terms),
                        incoterm=VALUES(incoterm),
                        address=VALUES(address)
                """, (
                    customer_name,
                    nd(str(row_dict.get('short_code') or '')),
                    nd(str(row_dict.get('country') or '')),
                    nd(str(row_dict.get('contact_person') or '')),
                    nd(str(row_dict.get('email') or '')),
                    nd(str(row_dict.get('phone') or '')),
                    nd(str(row_dict.get('payment_terms') or '')),
                    nd(str(row_dict.get('incoterm') or '')),
                    nd(str(row_dict.get('address') or '')),
                ))
                imported += 1
            except Exception as e:
                errors.append(f'Row {row_idx}: {str(e)}')
                skipped += 1

        db.commit()
        db.close()

        return jsonify({
            'success': True,
            'imported': imported,
            'skipped':  skipped,
            'errors':   errors,
            'message':  f'{imported} customers imported successfully'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@master_bp.route('/api/customers/template', methods=['GET'])
def download_customer_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    labels = ['Customer Name *','Short Code','Country','Contact Person','Email','Phone','Payment Terms','Inco Term','Address']
    header_fill = PatternFill('solid', start_color='1E4E8C')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

    for col, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    sample_fill = PatternFill('solid', start_color='EEF4FF')
    samples = [
        ['Wilo SE','WILO','Germany','Mr. Hans Mueller','hans@wilo.com','+49 231 4102 0','Payment against BL','CIF Hamburg','Nortkirchenstr. 100, Dortmund'],
        ['KSB SE','KSB','Germany','Mr. Peter Schmidt','peter@ksb.com','+49 6233 86 0','LC at sight','FOB Mumbai','Johann-Klein-Str. 9, Frankenthal'],
    ]
    for row_idx, row_data in enumerate(samples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = sample_fill
            cell.border = border

    col_widths = [30,15,15,25,30,20,30,20,40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 25

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='customer_import_template.xlsx', as_attachment=True)